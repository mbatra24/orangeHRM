import time

import openpyxl
from openpyxl import workbook
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.expected_conditions import element_to_be_clickable, presence_of_element_located
from selenium.webdriver.support.wait import WebDriverWait

from pages.pim_page import pim_page
from variables.variables import PIM_PAGE_URL


# employees_firstname = ["tester", "James", "Bond"]
# employees_lastname = ["1234", "bond", "rasuman"]

# employees_firstname = ["tester"]
# employees_lastname = ["1234"]

# reading data from an excel file

def read_data_from_excel():
    workbook = openpyxl.load_workbook(r"C:\Users\Mannu PC\PycharmProjects\orangeHRM_PYTEST\data\data.xlsx")
    sheet = workbook["Sheet1"]

    data = []
    for row in range(3, sheet.max_row + 1):
        firstname = sheet.cell(row, 1).value
        lastname = sheet.cell(row, 2).value
        data.append((firstname, lastname))

    return data


def add_employee(driver, data_output):
    # data_output = read_data_from_excel()
    # data_output = read_data_from_csv()
    # data_output= read_data_from_json()
    print(data_output)

    # below method to read file from excel
    pim_page(driver)
    driver.find_element(By.CSS_SELECTOR,"div.oxd-layout-context > div > div.orangehrm-paper-container > div.orangehrm-header-container > button").click()
    time.sleep(3)
    for firstname, lastname in data_output:
        driver.find_element(By.NAME, "firstName").send_keys(firstname)
        time.sleep(1)
        print("***********")
        print(driver.find_element(By.NAME, "firstName").get_attribute("value"))
        driver.find_element(By.NAME, "lastName").send_keys(lastname)
        time.sleep(2)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(3)
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/addEmployee")
        time.sleep(3)
        # pim_page(driver)
    driver.back()
    driver.refresh()
    time.sleep(2)
    print(driver.find_element(By.CSS_SELECTOR, "div.orangehrm-paper-container > div:nth-child(2) > div > span").text)
    time.sleep(2)

def del_employee(driver, data_output):
    for firstname, lastname in data_output:
        # empl name input field
        emp_name =driver.find_element(By.CSS_SELECTOR,"div > div:nth-child(1) > div > div:nth-child(2) > div > div > input")
        time.sleep(2)
        # sending first name last name
        emp_name.send_keys(firstname + Keys.SPACE + lastname)
        # emp_name.send_keys(lastname)
        time.sleep(2)
        emp_name.send_keys(Keys.ARROW_DOWN+Keys.RETURN)
        time.sleep(2)
        driver.save_screenshot("emp_search.png")
        # clicking search button
        driver.find_element(By.CSS_SELECTOR,"button.oxd-button.oxd-button--medium.oxd-button--secondary.orangehrm-left-space").click()
        time.sleep(4)
        driver.save_screenshot("search_results.png")
        # getting text from the search records
        print(driver.find_element(By.CSS_SELECTOR, "div.orangehrm-paper-container > div:nth-child(2) > div > span").text)
        time.sleep(2)
        print(driver.find_element(By.XPATH, "(//div[@role='cell'])[3]").text)
        time.sleep(2)
        print(driver.find_element(By.XPATH, "(//div[@role='cell'])[4]").text)
        delete = WebDriverWait(driver,10).until(element_to_be_clickable((By.CSS_SELECTOR, ".oxd-icon.bi-trash")))
        delete.click()
        del_text = WebDriverWait(driver, 10).until(presence_of_element_located((By.CSS_SELECTOR, ".oxd-text.oxd-text--p.oxd-text--card-body")))
        print(del_text.text)
        WebDriverWait(driver, 10).until(element_to_be_clickable((By.CSS_SELECTOR, ".oxd-icon.bi-trash.oxd-button-icon"))).click()
        # time.sleep(2)
        try:
            # driver.find_element(By.XPATH, "//p[contains(text(),'Success')]")
            print(WebDriverWait(driver, 10).until(presence_of_element_located((By.XPATH, "//p[contains(text(),'Success')]"))).text)
        except Exception:
            pass
        driver.save_screenshot("delete_success_new.png")
        # refresing driver
        driver.refresh()
        time.sleep(2)
        driver.save_screenshot("search_results_new.png")






# def add_employee(driver):
#     pim_page(driver)
#     driver.find_element(By.CSS_SELECTOR,"div.oxd-layout-context > div > div.orangehrm-paper-container > div.orangehrm-header-container > button").click()
#     time.sleep(3)
#     for first, last in zip(employees_firstname,employees_lastname):
#         driver.find_element(By.NAME, "firstName").send_keys(first)
#         # time.sleep(2)
#         driver.find_element(By.NAME, "lastName").send_keys(last)
#         time.sleep(2)
#         driver.find_element(By.XPATH, "//button[@type='submit']").click()
#         time.sleep(3)
#         driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/addEmployee")
#         time.sleep(3)





# This method searches for the employees added in the above function to assert
# that a successful search result is returned for each new employee
# def verify_employees_added(driver):
#     if driver.current_url != PIM_PAGE_URL:
#         driver.get(PIM_PAGE_URL)
#         time.sleep(3)
#
#     for first, last in zip(employees_firstname, employees_lastname):
#         driver.find_element(By.CSS_SELECTOR,"div > div:nth-child(1) > div > div:nth-child(2) > div > div > input").send_keys(f"{first} {last}")
#         driver.find_element(By.CSS_SELECTOR,"button.oxd-button.oxd-button--medium.oxd-button--secondary.orangehrm-left-space").click()
#         time.sleep(5)
#         print(driver.find_element(By.CSS_SELECTOR, "div.orangehrm-paper-container > div:nth-child(2) > div > span").text)
#         status = driver.find_element(By.CSS_SELECTOR,"div.orangehrm-paper-container > div:nth-child(2) > div > span").text
#         print(f"status = {status}")
#         assert status.find("(") >= 0
#         driver.get(PIM_PAGE_URL)
#         time.sleep(3)
